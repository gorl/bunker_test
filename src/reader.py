from typing import Dict

from openpyxl import load_workbook

from model import EntityContainer, Entity


def load_data(fname:str) -> Dict[str, EntityContainer]:
    result = {}
    wb = load_workbook(filename=fname)

    for type_name in wb.sheetnames:
        if type_name == 'Карты действий':
            # TODO ПОПРАВИТЬ ФОРМАТ!
            continue
        sheet = wb[type_name]
        entity_list = [Entity(name=name.value, score=int(score.value)) for name, score in zip(sheet['A'], sheet['C'])]

        c = int(sheet['F1'].value[1:])
        result[type_name] = EntityContainer(name=type_name, c=c, entity_list=entity_list)
    return result